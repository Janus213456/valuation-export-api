from io import BytesIO
from pathlib import Path
from typing import List, Optional
import base64
from urllib.parse import quote
from uuid import uuid4

from fastapi import FastAPI, HTTPException
from fastapi.responses import StreamingResponse, FileResponse
from pydantic import BaseModel
from openpyxl import load_workbook
from certificate_api import router as certificate_router

app = FastAPI(title="Valuation Export API")
app.include_router(certificate_router)

TEMPLATE_FILE = Path("Master_Valuation_Template_Clean_And_Final.xlsx")
GENERATED_DIR = Path("generated_files")
GENERATED_DIR.mkdir(exist_ok=True)


class ExportRow(BaseModel):
    suburb: Optional[str] = ""
    street_number: Optional[str] = ""
    street_name: str
    roof_m2: Optional[float] = None
    erf_m2: Optional[float] = None
    registration_date: str
    sale_price: float


class SubjectProperty(BaseModel):
    address: str
    type: str
    size_m2: float


class ExportValuationRequest(BaseModel):
    filename: str
    subject_property: SubjectProperty
    step2_rows: List[ExportRow] = []
    step3_rows: List[ExportRow] = []


def write_row(ws, excel_row: int, row: ExportRow) -> None:
    ws[f"B{excel_row}"] = row.suburb or ""
    ws[f"C{excel_row}"] = row.street_number or ""
    ws[f"D{excel_row}"] = row.street_name
    ws[f"E{excel_row}"] = row.roof_m2
    ws[f"F{excel_row}"] = row.erf_m2
    ws[f"G{excel_row}"] = row.registration_date
    ws[f"K{excel_row}"] = row.sale_price


def build_workbook_bytes(payload: ExportValuationRequest) -> tuple[bytes, str]:
    if not TEMPLATE_FILE.exists():
        raise HTTPException(status_code=500, detail="Template file not found on server.")

    try:
        wb = load_workbook(TEMPLATE_FILE)
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to load template: {str(e)}")

    ws["C4"] = payload.subject_property.address
    ws["D22"] = payload.subject_property.size_m2

    for idx, row in enumerate(payload.step2_rows[:3], start=8):
        write_row(ws, idx, row)

    for idx, row in enumerate(payload.step3_rows[:3], start=15):
        write_row(ws, idx, row)

    output = BytesIO()
    wb.save(output)
    file_bytes = output.getvalue()

    safe_filename = payload.filename if payload.filename.endswith(".xlsx") else f"{payload.filename}.xlsx"
    return file_bytes, safe_filename


@app.get("/")
def healthcheck():
    return {"status": "ok"}


@app.post("/export/valuation")
def export_valuation(payload: ExportValuationRequest):
    file_bytes, safe_filename = build_workbook_bytes(payload)

    return StreamingResponse(
        BytesIO(file_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe_filename}"'}
    )


@app.post("/export/valuation/gpt")
def export_valuation_gpt(payload: ExportValuationRequest):
    file_bytes, safe_filename = build_workbook_bytes(payload)
    base64_content = base64.b64encode(file_bytes).decode("utf-8")

    return {
        "openaiFileResponse": [
            {
                "name": safe_filename,
                "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "content": base64_content
            }
        ]
    }


@app.post("/export/valuation/gpt-link")
def export_valuation_gpt_link(payload: ExportValuationRequest):
    file_bytes, safe_filename = build_workbook_bytes(payload)

    unique_name = f"{uuid4()}_{safe_filename}"
    file_path = GENERATED_DIR / unique_name
    file_path.write_bytes(file_bytes)

    encoded_name = quote(unique_name)
    download_url = f"https://valuation-export-api.onrender.com/download/{encoded_name}"

    return {
        "openaiFileResponse": [
            {
                "name": safe_filename,
                "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "download_link": download_url
            }
        ]
    }


@app.get("/download/{file_name}")
def download_generated_file(file_name: str):
    file_path = GENERATED_DIR / file_name

    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found.")

    original_name = file_name.split("_", 1)[1] if "_" in file_name else file_name

    return FileResponse(
        path=file_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=original_name
    )